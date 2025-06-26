# B&R Capital Excel Data Extraction System

# Code Structure Summary:
# 
# 1. Cell Mapping Parser (lines 13-90)
#    - Reads the Excel reference file containing ~1,200 mappings
#    - Validates required columns exist
#    - Creates structured dictionary of mappings
#    - Handles missing data gracefully
#
# 2. Excel Data Extractor (lines 92-191)
#    - Supports both .xlsb and .xlsx/xlsm files
#    - Extracts values based on cell mappings
#    - Handles formula errors and missing values
#    - Returns structured data with metadata
#
# 3. Batch File Processor (lines 193-285)
#    - Processes multiple files in configurable batches
#    - Tracks progress and performance
#    - Generates summary reports
#    - Handles failures without stopping
#
# 4. Error Handling Utilities (lines 287-338)
#    - Custom exception classes
#    - Comprehensive logging
#    - Error recovery mechanisms
#    - NaN handling for missing values
#
# 5. Main Processing Pipeline (lines 340-end)
#    - Orchestrates the extraction process
#    - Provides command-line interface
#    - Generates output reports

"""
B&R Capital Dashboard - Excel Data Extraction System
Extracts ~1,200 data points from underwriting models
"""

import os
import sys
import json
import logging
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
import pyxlsb
import structlog

# Import the comprehensive error handling system
from .error_handling_system import ErrorHandler, ErrorCategory, process_cell_value_with_error_handling
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed
import warnings

# Configure structured logging
structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer()
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    cache_logger_on_first_use=True,
)

logger = structlog.get_logger()

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


@dataclass
class CellMapping:
    """Represents a single cell mapping from the reference file"""
    category: str
    description: str
    sheet_name: str
    cell_address: str
    field_name: str  # Cleaned description for database field


class CellMappingParser:
    """Parses the Excel reference file containing ~1,200 cell mappings"""
    
    def __init__(self, reference_file_path: str):
        self.reference_file_path = reference_file_path
        self.mappings: Dict[str, CellMapping] = {}
        self.logger = structlog.get_logger().bind(
            component="CellMappingParser"
        )
        
    def load_mappings(self) -> Dict[str, CellMapping]:
        """Load and parse cell mappings from reference Excel file"""
        try:
            self.logger.info("loading_cell_mappings", 
                             file_path=self.reference_file_path)
            
            # Read the reference file
            df = pd.read_excel(
                self.reference_file_path,
                sheet_name="UW Model - Cell Reference Table"
            )

            # Debug: Print column info
            print(f"Columns in dataframe: {list(df.columns)}")
            print(f"DataFrame shape: {df.shape}")
            
            # Map to standard column references based on actual columns
            # Looking at the output, we need to identify the correct columns
            col_mapping = {}
            
            # Try to identify columns by content/name
            for i, col_name in enumerate(df.columns):
                col_letter = chr(65 + i)  # A, B, C, D, etc.
                col_mapping[col_letter] = col_name
            
            # Use the identified columns
            category_col = col_mapping.get('B', df.columns[1] if len(df.columns) > 1 else None)
            description_col = col_mapping.get('C', df.columns[2] if len(df.columns) > 2 else None) 
            sheet_col = col_mapping.get('D', df.columns[3] if len(df.columns) > 3 else None)
            cell_col = col_mapping.get('G', df.columns[6] if len(df.columns) > 6 else None)
            
            print(f"Using columns - Category: {category_col}, Description: {description_col}, Sheet: {sheet_col}, Cell: {cell_col}")
            
            # Process each row
            mapping_count = 0
            for idx, row in df.iterrows():
                if pd.notna(row[description_col]) and pd.notna(row[cell_col]):
                    # Clean field name from description
                    field_name = self._clean_field_name(str(row[description_col]))
                    
                    mapping = CellMapping(
                        category=str(row[category_col]) if pd.notna(row[category_col]) else "Uncategorized",
                        description=str(row[description_col]),
                        sheet_name=str(row[sheet_col]),
                        cell_address=str(row[cell_col]).strip().upper(),
                        field_name=field_name
                    )
                    
                    self.mappings[field_name] = mapping
                    mapping_count += 1
            
            self.logger.info("mappings_loaded", 
                           count=mapping_count,
                           categories=len(set(m.category for m in self.mappings.values())))
            
            return self.mappings
            
        except Exception as e:
            self.logger.error("failed_to_load_mappings", 
                            error=str(e),
                            file_path=self.reference_file_path)
            raise
    
    def _clean_field_name(self, description: str) -> str:
        """Convert description to clean field name"""
        # Remove special characters and spaces
        clean_name = description.strip()
        clean_name = clean_name.replace(" ", "_")
        clean_name = clean_name.replace("-", "_")
        clean_name = clean_name.replace("(", "")
        clean_name = clean_name.replace(")", "")
        clean_name = clean_name.replace("/", "_")
        clean_name = clean_name.replace(".", "")
        clean_name = clean_name.upper()
        
        # Remove duplicate underscores
        while "__" in clean_name:
            clean_name = clean_name.replace("__", "_")
            
        return clean_name
    
    def export_mapping_summary(self, output_path: str):
        """Export mapping summary for documentation"""
        # Ensure directory exists
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)
        
        summary_data = []
        for field_name, mapping in self.mappings.items():
            summary_data.append({
                'Field Name': field_name,
                'Category': mapping.category,
                'Description': mapping.description,
                'Sheet': mapping.sheet_name,
                'Cell': mapping.cell_address
            })
        
        df = pd.DataFrame(summary_data)
        df.to_csv(output_path, index=False)
        self.logger.info("mapping_summary_exported", path=output_path)


class ExcelDataExtractor:
    """Extracts data from Excel underwriting models (.xlsb and .xlsx)"""
    
    def __init__(self, cell_mappings: Dict[str, CellMapping]):
        self.mappings = cell_mappings
        self.logger = structlog.get_logger().bind(
            component="ExcelDataExtractor"
        )
        # Initialize comprehensive error handling
        self.error_handler = ErrorHandler()
        
    def extract_from_file(self, file_path: str, file_content: bytes = None) -> Dict[str, Any]:
        """Extract all mapped values from an Excel file"""
        start_time = datetime.now()
        extracted_data = {
            '_file_path': file_path,
            '_extraction_timestamp': datetime.now().isoformat(),
            '_extraction_errors': []
        }
        
        try:
            # Validate file exists when using file path
            if file_content is None and not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
                
            # Determine file type
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.xlsb':
                workbook = self._load_xlsb(file_path, file_content)
            else:
                workbook = self._load_xlsx(file_path, file_content)
            
            # First, let's check what sheets are available
            if hasattr(workbook, 'sheets'):  # pyxlsb
                available_sheets = list(workbook.sheets)
                print(f"Available sheets in xlsb file: {available_sheets}")
            else:  # openpyxl
                available_sheets = workbook.sheetnames
                print(f"Available sheets in xlsx file: {available_sheets}")
            
            # Extract values for each mapping
            successful_extractions = 0
            failed_extractions = 0
            debug_count = 0
            
            for field_name, mapping in self.mappings.items():
                try:
                    value = self._extract_cell_value(
                        workbook, 
                        mapping.sheet_name, 
                        mapping.cell_address,
                        field_name
                    )
                    extracted_data[field_name] = value
                    successful_extractions += 1
                    
                    # Debug first few extractions
                    if debug_count < 10:
                        print(f"DEBUG: {field_name} = {value} (from {mapping.sheet_name}!{mapping.cell_address})")
                        debug_count += 1
                    
                except Exception as e:
                    # Log error but continue extraction
                    extracted_data[field_name] = np.nan
                    extracted_data['_extraction_errors'].append({
                        'field': field_name,
                        'sheet': mapping.sheet_name,
                        'cell': mapping.cell_address,
                        'error': str(e)
                    })
                    failed_extractions += 1
                    
                    # Debug first few errors
                    if debug_count < 10:
                        print(f"ERROR: {field_name} from {mapping.sheet_name}!{mapping.cell_address}: {str(e)}")
                        debug_count += 1
            
            # Add extraction metadata including comprehensive error report
            duration = (datetime.now() - start_time).total_seconds()
            error_summary = self.error_handler.get_error_summary()
            
            extracted_data['_extraction_metadata'] = {
                'total_fields': len(self.mappings),
                'successful': successful_extractions,
                'failed': failed_extractions,
                'duration_seconds': duration,
                'error_summary': error_summary
            }
            
            self.logger.info("file_extraction_complete",
                           file_path=file_path,
                           successful=successful_extractions,
                           failed=failed_extractions,
                           duration=duration)
            
            return extracted_data
            
        except Exception as e:
            self.logger.error("file_extraction_failed",
                            file_path=file_path,
                            error=str(e))
            raise
    
    def _load_xlsb(self, file_path: str, file_content: bytes = None):
        """Load .xlsb file using pyxlsb"""
        if file_content:
            # Load from bytes (e.g., from SharePoint)
            import io
            return pyxlsb.open_workbook(io.BytesIO(file_content))
        else:
            # Load from file path
            return pyxlsb.open_workbook(file_path)
    
    def _load_xlsx(self, file_path: str, file_content: bytes = None):
        """Load .xlsx/.xlsm file"""
        if file_content:
            import io
            return openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=True,  # Get calculated values, not formulas
                keep_vba=True    # Preserve VBA for .xlsm files
            )
        else:
            return openpyxl.load_workbook(
                file_path,
                data_only=True,
                keep_vba=True
            )
    
    def _extract_cell_value(self, workbook, sheet_name: str, cell_address: str, field_name: str = "") -> Any:
        """Extract value from specific cell with comprehensive error handling"""
        try:
            # Check if it's a pyxlsb workbook
            if hasattr(workbook, 'sheets'):  # pyxlsb
                if sheet_name not in workbook.sheets:
                    available_sheets = list(workbook.sheets)
                    return self.error_handler.handle_missing_sheet(field_name, sheet_name, available_sheets)
                
                # Clean cell address - remove $ signs if present
                clean_address = cell_address.replace('$', '').upper()
                
                # Parse cell address (e.g., "A1" -> row=1, col=1)
                import re
                match = re.match(r'^([A-Z]+)(\d+)$', clean_address)
                if not match:
                    return self.error_handler.handle_invalid_cell_address(
                        field_name, sheet_name, cell_address, "Invalid format - expected format like 'A1', 'B10'"
                    )
                
                col_str, row_str = match.groups()
                # Convert from Excel 1-based to pyxlsb 0-based indexing
                # IMPORTANT: pyxlsb uses 0-based indexing while Excel uses 1-based
                # Excel D6 (row=6, col=4) → pyxlsb (row=5, col=3)
                # See docs/PYXLSB_INDEXING_GUIDE.md for detailed explanation
                target_row = int(row_str) - 1  # Convert to 0-based row
                
                # Convert column letters to number (A=1, B=2, AA=27, etc.) then to 0-based
                target_col = 0
                for char in col_str:
                    target_col = target_col * 26 + (ord(char) - ord('A') + 1)
                target_col = target_col - 1  # Convert to 0-based column
                
                # Debug first few cell reads
                debug_info = f"Looking for {sheet_name}!{clean_address} (row={target_row}, col={target_col})"
                
                # Get the sheet and read all rows to find our target
                with workbook.get_sheet(sheet_name) as sheet:
                    found_value = None
                    rows_checked = 0
                    
                    # Read all rows in the sheet
                    for row_data in sheet.rows():
                        if not row_data:  # Skip empty rows
                            continue
                            
                        rows_checked += 1
                        
                        # Check if this is our target row (pyxlsb uses 0-based indexing)
                        for cell in row_data:
                            if cell.r == target_row and cell.c == target_col:
                                found_value = cell.v
                                # Debug successful finds for first few
                                if rows_checked <= 20:  # Only debug first 20 rows
                                    print(f"FOUND: {debug_info} = {found_value}")
                                return self.error_handler.process_cell_value(
                                    found_value, field_name, sheet_name, cell_address
                                )
                    
                    # If we get here, cell wasn't found - try different approach
                    # Some xlsb files might have sparse data, so let's try reading the specific cell directly
                    with workbook.get_sheet(sheet_name) as sheet:
                        # Alternative method: iterate through all cells to find our target
                        for row in sheet.rows():
                            for cell in row:
                                if hasattr(cell, 'r') and hasattr(cell, 'c'):
                                    if cell.r == target_row and cell.c == target_col:
                                        return self.error_handler.process_cell_value(
                                            cell.v, field_name, sheet_name, cell_address
                                        )
                    
                    # Still not found - cell is likely empty or outside bounds
                    if rows_checked <= 10:  # Debug for first few attempts
                        print(f"NOT FOUND: {debug_info} (checked {rows_checked} rows)")
                    return self.error_handler.handle_cell_not_found(
                        field_name, sheet_name, cell_address, (rows_checked, target_col + 1)
                    )
                
            else:  # openpyxl
                if sheet_name not in workbook.sheetnames:
                    available_sheets = workbook.sheetnames
                    return self.error_handler.handle_missing_sheet(field_name, sheet_name, available_sheets)
                    
                sheet = workbook[sheet_name]
                
                # Get cell value
                cell = sheet[cell_address]
                return self.error_handler.process_cell_value(
                    cell.value, field_name, sheet_name, cell_address
                )
                
        except Exception as e:
            # Use error handler for unknown errors
            return self.error_handler.handle_unknown_error(
                field_name, sheet_name, cell_address, str(e)
            )
    
    def _process_cell_value(self, value: Any) -> Any:
        """Process and clean cell values"""
        # Handle None/empty
        if value is None or value == '':
            return np.nan
            
        # Handle error values
        if isinstance(value, str):
            error_indicators = ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', 
                              '#N/A', '#NULL!', '#NUM!']
            if any(err in str(value) for err in error_indicators):
                return np.nan
        
        # Return cleaned value
        return value


class BatchFileProcessor:
    """Processes multiple Excel files in batches"""
    
    def __init__(self, extractor: ExcelDataExtractor, batch_size: int = 10):
        self.extractor = extractor
        self.batch_size = batch_size
        self.logger = structlog.get_logger().bind(
            component="BatchFileProcessor"
        )
        
    def process_files(self, file_list: List[Dict[str, Any]], 
                     max_workers: int = 4) -> List[Dict[str, Any]]:
        """Process multiple files in batches with parallel execution"""
        total_files = len(file_list)
        processed_results = []
        failed_files = []
        
        self.logger.info("starting_batch_processing",
                        total_files=total_files,
                        batch_size=self.batch_size,
                        max_workers=max_workers)
        
        # Process in batches
        for batch_start in range(0, total_files, self.batch_size):
            batch_end = min(batch_start + self.batch_size, total_files)
            batch = file_list[batch_start:batch_end]
            
            self.logger.info("processing_batch",
                           batch_number=batch_start // self.batch_size + 1,
                           batch_size=len(batch))
            
            # Process batch with thread pool
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit extraction tasks
                future_to_file = {
                    executor.submit(
                        self._process_single_file, 
                        file_info
                    ): file_info 
                    for file_info in batch
                }
                
                # Collect results
                for future in as_completed(future_to_file):
                    file_info = future_to_file[future]
                    try:
                        result = future.result()
                        processed_results.append(result)
                    except Exception as e:
                        self.logger.error("file_processing_failed",
                                        file_path=file_info.get('file_path'),
                                        error=str(e))
                        failed_files.append({
                            'file_info': file_info,
                            'error': str(e)
                        })
        
        # Generate summary
        self._generate_processing_summary(processed_results, failed_files)
        
        return processed_results
    
    def _process_single_file(self, file_info: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single file"""
        file_path = file_info.get('file_path')
        file_content = file_info.get('file_content')  # If already downloaded
        
        # Extract data
        extracted_data = self.extractor.extract_from_file(
            file_path, 
            file_content
        )
        
        # Add file metadata
        extracted_data.update({
            '_deal_name': file_info.get('deal_name'),
            '_deal_stage': file_info.get('deal_stage'),
            '_file_modified_date': file_info.get('modified_date')
        })
        
        return extracted_data
    
    def _generate_processing_summary(self, results: List[Dict], 
                                   failed_files: List[Dict]):
        """Generate summary report of batch processing"""
        summary = {
            'total_processed': len(results),
            'total_failed': len(failed_files),
            'total_fields_extracted': sum(
                r['_extraction_metadata']['successful'] 
                for r in results
            ),
            'average_extraction_time': np.mean([
                r['_extraction_metadata']['duration_seconds'] 
                for r in results
            ]) if results else 0,
            'failed_files': [f['file_info']['file_path'] for f in failed_files]
        }
        
        self.logger.info("batch_processing_complete", **summary)
        
        # Save summary to file
        summary_path = f"extraction_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Ensure directory exists if path includes directory
        if os.path.dirname(summary_path):
            os.makedirs(os.path.dirname(summary_path), exist_ok=True)
            
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)


# Custom Exception Classes
class ExtractionError(Exception):
    """Base exception for extraction errors"""
    pass

class FileAccessError(ExtractionError):
    """File access errors"""
    pass

class MappingError(ExtractionError):
    """Cell mapping errors"""
    pass

class DataValidationError(ExtractionError):
    """Data validation errors"""
    pass


# Utility Functions
def setup_logging(log_level: str = "INFO", log_file: str = None):
    """Set up logging configuration"""
    logging.basicConfig(
        level=getattr(logging, log_level),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_file) if log_file else logging.NullHandler()
        ]
    )

def validate_extracted_data(data: Dict[str, Any], 
                          required_fields: List[str]) -> Tuple[bool, List[str]]:
    """Validate extracted data has required fields"""
    missing_fields = []
    
    for field in required_fields:
        if field not in data or pd.isna(data[field]):
            missing_fields.append(field)
    
    is_valid = len(missing_fields) == 0
    return is_valid, missing_fields

def export_to_csv(results: List[Dict[str, Any]], output_path: str):
    """Export extraction results to CSV"""
    # Ensure directory exists
    output_dir = Path(output_path).parent
    if output_dir != Path('.'):  # Only create if not current directory
        output_dir.mkdir(parents=True, exist_ok=True)
    
    # Convert list of dictionaries to DataFrame
    df = pd.DataFrame(results)
    
    # Reorder columns to put metadata first
    metadata_cols = [col for col in df.columns if col.startswith('_')]
    data_cols = [col for col in df.columns if not col.startswith('_')]
    df = df[metadata_cols + sorted(data_cols)]
    
    # Export to CSV
    df.to_csv(output_path, index=False)
    logger.info("results_exported", path=output_path, records=len(df))


# Main Processing Pipeline
def main():
    """Main entry point for Excel extraction"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Extract data from B&R Capital underwriting models"
    )
    parser.add_argument(
        "--reference-file", 
        required=True,
        help="Path to Excel reference file with cell mappings"
    )
    parser.add_argument(
        "--file-list", 
        required=True,
        help="JSON file containing list of Excel files to process"
    )
    parser.add_argument(
        "--output-dir", 
        default="./extraction_output",
        help="Directory for output files"
    )
    parser.add_argument(
        "--batch-size", 
        type=int,
        default=10,
        help="Number of files to process in each batch"
    )
    parser.add_argument(
        "--max-workers", 
        type=int,
        default=4,
        help="Maximum number of parallel workers"
    )
    
    args = parser.parse_args()
    
    # Set up logging
    setup_logging(log_file="extraction.log")
    logger.info("starting_extraction_pipeline", **vars(args))
    
    try:
        # Create output directory
        os.makedirs(args.output_dir, exist_ok=True)
        
        # Load cell mappings
        logger.info("loading_cell_mappings")
        parser = CellMappingParser(args.reference_file)
        mappings = parser.load_mappings()
        
        # Export mapping summary
        parser.export_mapping_summary(
            os.path.join(args.output_dir, "mapping_summary.csv")
        )
        
        # Load file list
        logger.info("loading_file_list")
        with open(args.file_list, 'r') as f:
            file_list = json.load(f)
        
        # Create extractor and processor
        extractor = ExcelDataExtractor(mappings)
        processor = BatchFileProcessor(extractor, args.batch_size)
        
        # Process files
        logger.info("starting_file_processing", file_count=len(file_list))
        results = processor.process_files(file_list, args.max_workers)
        
        # Export results
        output_path = os.path.join(
            args.output_dir, 
            f"extraction_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        export_to_csv(results, output_path)
        
        logger.info("extraction_complete", 
                   files_processed=len(results),
                   output_path=output_path)
        
    except Exception as e:
        logger.error("extraction_pipeline_failed", error=str(e))
        raise


if __name__ == "__main__":
    main()